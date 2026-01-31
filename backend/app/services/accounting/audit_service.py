"""
Reconciliation Audit Service

This module provides services for managing reconciliation audit logs.
"""
from datetime import datetime
from typing import List, Optional, Dict, Any, Tuple

from ...models.reconciliation import (
from ...models.user import User
from ...schemas.reconciliation import (
from .base import BaseReconciliationService
from sqlalchemy import and_, or_, func, select, desc, asc
from sqlalchemy.orm import Session, joinedload
from uuid import UUID

from app.core.exceptions import (
from app.core.logging import get_logger



    NotFoundException,
    ValidationException,
    ForbiddenException
)

    ReconciliationAuditLog,
    Reconciliation
)
    ReconciliationAuditLogCreate
)

logger = get_logger(__name__)


class ReconciliationAuditService(BaseReconciliationService):
    """Service for handling reconciliation audit operations."""
    
    def get_audit_log(self, log_id: UUID, user_id: UUID) -> ReconciliationAuditLog:
        """Get Audit Log."""
        """Get an audit log entry by ID.
        
        Args:
            log_id: ID of the audit log entry to retrieve
            user_id: ID of the user making the request
            
        Returns:
            The audit log entry with user details
            
        Raises:
            NotFoundException: If the audit log entry doesn't exist
            ForbiddenException: If the user doesn't have permission to view this log entry
        """
        log = (
            self.db.query(ReconciliationAuditLog)
            .options(joinedload(ReconciliationAuditLog.user))
            .filter(ReconciliationAuditLog.id == log_id)
            .first()
        )
        
        if not log:
            raise NotFoundException("Audit log entry not found")
        
        # Check if the user has permission to view this log entry
        # Only the user who created the reconciliation or an admin can view the logs
        reconciliation = self.db.query(Reconciliation).get(log.reconciliation_id)
        if not reconciliation or (reconciliation.created_by != user_id and not self._is_admin(user_id)):
            raise ForbiddenException("You don't have permission to view this audit log")
        
        return log
    
    def list_audit_logs(
        self,
        reconciliation_id: UUID,
        user_id: UUID,
        action: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_filter: Optional[UUID] = None,
        skip: int = 0,
        limit: int = 100,
        sort_by: str = "created_at",
        sort_order: str = "desc"
    ) -> Tuple[List[ReconciliationAuditLog], int]:
        """List Audit Logs."""
        """List audit logs for a reconciliation with optional filtering and sorting.
        
        Args:
            reconciliation_id: ID of the reconciliation
            user_id: ID of the user making the request
            action: Filter by action type
            start_date: Filter by start date (greater than or equal)
            end_date: Filter by end date (less than or equal)
            user_filter: Filter by user ID who performed the action
            skip: Number of records to skip for pagination
            limit: Maximum number of records to return
            sort_by: Field to sort by (default: created_at)
            sort_order: Sort order (asc or desc, default: desc)
            
        Returns:
            Tuple of (list of audit logs, total count)
            
        Raises:
            NotFoundException: If the reconciliation doesn't exist
            ForbiddenException: If the user doesn't have permission to view the logs
            ValidationException: If the sort field or order is invalid
        """
        # Check if the reconciliation exists and the user has permission to view it
        reconciliation = self.db.query(Reconciliation).get(reconciliation_id)
        if not reconciliation:
            raise NotFoundException("Reconciliation not found")
        
        # Only the user who created the reconciliation or an admin can view the logs
        if reconciliation.created_by != user_id and not self._is_admin(user_id):
            raise ForbiddenException("You don't have permission to view these audit logs")
        
        # Validate sort field
        valid_sort_fields = ["created_at", "action", "user_id"]
        if sort_by not in valid_sort_fields:
            raise ValidationException(f"Invalid sort field. Must be one of: {', '.join(valid_sort_fields)}")
        
        # Validate sort order
        sort_order = sort_order.lower()
        if sort_order not in ["asc", "desc"]:
            raise ValidationException("Invalid sort order. Must be 'asc' or 'desc'")
        
        # Build the query
        query = (
            self.db.query(ReconciliationAuditLog)
            .options(joinedload(ReconciliationAuditLog.user))
            .filter(ReconciliationAuditLog.reconciliation_id == reconciliation_id)
        )
        
        # Apply filters
        if action:
            query = query.filter(ReconciliationAuditLog.action == action)
            
        if start_date:
            query = query.filter(ReconciliationAuditLog.created_at >= start_date)
            
        if end_date:
            # Add one day to include the entire end date
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(ReconciliationAuditLog.created_at <= end_date)
            
        if user_filter:
            query = query.filter(ReconciliationAuditLog.user_id == user_filter)
        
        # Apply sorting
        sort_column = getattr(ReconciliationAuditLog, sort_by, None)
        if sort_column is None:
            sort_column = ReconciliationAuditLog.created_at
        
        if sort_order == "asc":
            query = query.order_by(asc(sort_column))
        else:
            query = query.order_by(desc(sort_column))
        
        # Get total count
        total = query.count()
        
        # Apply pagination
        logs = query.offset(skip).limit(limit).all()
        
        return logs, total
    
    def create_audit_log(
        self, 
        reconciliation_id: UUID, 
        action: str, 
        user_id: UUID, 
        details: Optional[Dict[str, Any]] = None
    ) -> ReconciliationAuditLog:
        """Create Audit Log."""
        """Create a new audit log entry.
        
        Args:
            reconciliation_id: ID of the reconciliation
            action: Action that was performed
            user_id: ID of the user who performed the action
            details: Additional details about the action
            
        Returns:
            The created audit log entry
            
        Raises:
            NotFoundException: If the reconciliation or user doesn't exist
            ForbiddenException: If the user doesn't have permission to create a log entry
        """
        # Check if the reconciliation exists
        reconciliation = self.db.query(Reconciliation).get(reconciliation_id)
        if not reconciliation:
            raise NotFoundException("Reconciliation not found")
        
        # Check if the user exists
        user = self.db.query(User).get(user_id)
        if not user:
            raise NotFoundException("User not found")
        
        # Only the user who created the reconciliation or an admin can add audit logs
        if reconciliation.created_by != user_id and not self._is_admin(user_id):
            raise ForbiddenException("You don't have permission to add audit logs for this reconciliation")
        
        # Create the audit log
        log = ReconciliationAuditLog(
            id=uuid4(),
            reconciliation_id=reconciliation_id,
            action=action,
            details=details or {},
            user_id=user_id,
            created_at=datetime.utcnow()
        )
        
        self.db.add(log)
        self.db.commit()
        
        # Refresh to load relationships
        self.db.refresh(log)
        
        return log
    
    def export_audit_logs(
        self, 
        reconciliation_id: UUID, 
        user_id: UUID,
        format: str = "csv"
    ) -> str:
        """Export Audit Logs."""
        """Export audit logs for a reconciliation in the specified format.
        
        Args:
            reconciliation_id: ID of the reconciliation
            user_id: ID of the user making the request
            format: Export format (csv, json, xlsx)
            
        Returns:
            The exported audit logs as a string or file path
            
        Raises:
            NotFoundException: If the reconciliation doesn't exist
            ForbiddenException: If the user doesn't have permission to export the logs
            ValidationException: If the format is invalid
        """
        # Check if the reconciliation exists and the user has permission to view it
        reconciliation = self.db.query(Reconciliation).get(reconciliation_id)
        if not reconciliation:
            raise NotFoundException("Reconciliation not found")
        
        # Only the user who created the reconciliation or an admin can export the logs
        if reconciliation.created_by != user_id and not self._is_admin(user_id):
            raise ForbiddenException("You don't have permission to export these audit logs")
        
        # Validate format
        format = format.lower()
        if format not in ["csv", "json", "xlsx"]:
            raise ValidationException("Invalid format. Must be one of: csv, json, xlsx")
        
        # Get all audit logs for the reconciliation
        logs, _ = self.list_audit_logs(
            reconciliation_id=reconciliation_id,
            user_id=user_id,
            skip=0,
            limit=10000  # Maximum number of logs to export
        )
        
        # Convert logs to the requested format
        if format == "csv":
            return self._export_to_csv(logs)
        elif format == "json":
            return self._export_to_json(logs)
        elif format == "xlsx":
            return self._export_to_xlsx(logs)
    
    def _export_to_csv(self, logs: List[ReconciliationAuditLog]) -> str:
        """ Export To Csv."""
        """Export audit logs to CSV format.
        
        Args:
            logs: List of audit log entries
            
        Returns:
            CSV string
        """
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "Timestamp",
            "Action",
            "User",
            "Email",
            "Details"
        ])
        
        # Write data
        for log in logs:
            writer.writerow([
                log.created_at.isoformat(),
                log.action,
                f"{log.user.first_name} {log.user.last_name}" if log.user else "System",
                log.user.email if log.user else "system@paksa.com",
                str(log.details)
            ])
        
        return output.getvalue()
    
    def _export_to_json(self, logs: List[ReconciliationAuditLog]) -> str:
        """ Export To Json."""
        """Export audit logs to JSON format.
        
        Args:
            logs: List of audit log entries
            
        Returns:
            JSON string
        """
        import json
        
        log_data = []
        
        for log in logs:
            log_data.append({
                "timestamp": log.created_at.isoformat(),
                "action": log.action,
                "user": {
                    "id": str(log.user.id) if log.user else None,
                    "name": f"{log.user.first_name} {log.user.last_name}" if log.user else "System",
                    "email": log.user.email if log.user else "system@paksa.com"
                },
                "details": log.details
            })
        
        return json.dumps(log_data, indent=2)
    
    def _export_to_xlsx(self, logs: List[ReconciliationAuditLog]) -> str:
        """ Export To Xlsx."""
        """Export audit logs to Excel format.
        
        Args:
            logs: List of audit log entries
            
        Returns:
            Path to the generated Excel file
        """
        import os
        import tempfile
        from openpyxl import Workbook
        
        # Create a temporary file
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        # Write header
        headers = ["Timestamp", "Action", "User", "Email", "Details"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Write data
        for row_num, log in enumerate(logs, 2):
            ws.cell(row=row_num, column=1, value=log.created_at.isoformat())
            ws.cell(row=row_num, column=2, value=log.action)
            ws.cell(row=row_num, column=3, value=f"{log.user.first_name} {log.user.last_name}" if log.user else "System")
            ws.cell(row=row_num, column=4, value=log.user.email if log.user else "system@paksa.com")
            ws.cell(row=row_num, column=5, value=str(log.details))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Save the workbook
        wb.save(temp_path)
        
        return temp_path
    
    def _is_admin(self, user_id: UUID) -> bool:
        """ Is Admin."""
        """Check if a user is an admin.
        
        Args:
            user_id: ID of the user to check
            
        Returns:
            True if the user is an admin, False otherwise
        """
        # This is a placeholder implementation
        return False
